from decimal import Decimal
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font


def myround(x, w):
    return float(Decimal(str(x)).quantize(Decimal("0." + "0" * (w - 1) + "1"), rounding="ROUND_HALF_UP"))


def merge_df_cells(path, cells):
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active
    side = Side(border_style='thin', color='000000')
    for i in range(1, worksheet.max_row + 1):
        for j in range(1, worksheet.max_column + 1):
            worksheet.cell(i, j).alignment = Alignment(horizontal='center', vertical='center')
            worksheet.cell(i, j).border = Border(top=side, bottom=side, left=side, right=side)
            worksheet.cell(i, j).font = Font(bold=False)

    for cell in cells:
        worksheet.merge_cells(cell)
    workbook.save(path)
